import warnings
import joblib
import numpy as np
import pandas as pd
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import MinMaxScaler
from sklearn.utils import resample
import openpyxl
from scipy import stats

warnings.filterwarnings("ignore")
rng = np.random.RandomState(0)
np.set_printoptions(threshold=np.inf)
np.set_printoptions(suppress=True)

def data_read():
    data = pd.read_excel(r'data_reg-1114.xlsx')
    x0 = data.iloc[:, 2:-4]
    x1 = data.iloc[:, 2:-3]
    y_Tem = data.iloc[:, -4]
    y_Wloss = data.iloc[:, -3]

    y_Tem = np.ravel(y_Tem)  # 降维
    y_Wloss = np.ravel(y_Wloss)  # 降维
    y_Tem.astype('float')
    y_Wloss.astype('float')
    # x, y = shuffle(x, y, random_state=10)  # 打乱顺序
    return x0, x1, y_Tem, y_Wloss


def model_pre(x, y, name, path, params,space_pre):
    params = np.load(params, allow_pickle=True)

    model = joblib.load(path)

    X_ = x
    y_ = y

    mm_y = MinMaxScaler()
    for param in params:
        if param['type'] == name:
            m = param
            if 'normalize' in param:
                if param['normalize'] == 1:
                    mm_x = MinMaxScaler()
                    X_ = mm_x.fit_transform(X_)
                    space_pre = mm_x.fit_transform(space_pre)

            if 'scale' in param:
                if param['scale'] == 1:
                    ss_y = StandardScaler()
                    X_ = ss_y.fit_transform(X_)
                    space_pre = ss_y.fit_transform(space_pre)

            if 'normalize_y' in param:
                if param['normalize_y'] == 1:
                    y_ = mm_y.fit_transform(y_.reshape(-1, 1))
                    y_ = np.ravel(y_)
                    y_.astype(int)

    model.fit(X_, y_)

    y_pre = model.predict(space_pre)

    if m['normalize_y'] == 1:
        y_pre = mm_y.inverse_transform(y_pre.reshape(-1, 1))

    return y_pre


def PA(goal, searchspace0):
    p = (goal - searchspace0[:, 0]) / (searchspace0[:, 1])
    pas = 1 - stats.norm.cdf(p)
    return (pas)


def data_write(data, path):
    outwb = openpyxl.Workbook()
    ws = outwb.create_sheet(index=1)
    i = 1
    r = 1
    for line in data:
        for col in range(1, len(line) + 1):
            # ColNum = r
            ws.cell(row=r, column=col).value = line[col - 1]
        i += 1
        r += 1
    savexlsx = path
    outwb.save(savexlsx)


if __name__ == '__main__':
    stand_scaler = StandardScaler()
    norm_scaler = MinMaxScaler()

    space_pre = pd.read_excel('07-预测空间.xlsx')

    x0, x1, y_Tem, y_Wloss = data_read()

    # Tem预测
    pre_a = np.arange(0, len(space_pre))
    n = len(y_Tem)
    m = 1000

    for u in range(0, m):
        x11, y11 = resample(x0, y_Tem, replace=True, n_samples=n, random_state=u)
        predict_y = model_pre(x11, y11, 'xgboost',r'G:\Grade one\9-Dataset\07-Ce-TGA\04-建模\07-回归任务-1114\02-model\xgboost.pkl'
                                ,r'G:\Grade one\9-Dataset\07-Ce-TGA\04-建模\07-回归任务-1114\02-model\03-models_params-6D_copy.npy'
                                , space_pre)
        print(u)
        pre_a = np.column_stack((pre_a, predict_y))

    space_pre['pre_tem'] = np.mean(pre_a[:, 1:, ], axis=1)

    # Def的PA计算
    pre_a = np.arange(0, len(space_pre))
    n = len(y_Wloss)
    m = 1000

    for u in range(0, m):
        x11, y11 = resample(x1, y_Wloss, replace=True, n_samples=n, random_state=u)
        predict_y = model_pre(x11, y11, 'xgboost',r'G:\Grade one\9-Dataset\07-Ce-TGA\04-建模\07-回归任务-1114\03-model\xgboost.pkl'
                                 ,r'G:\Grade one\9-Dataset\07-Ce-TGA\04-建模\07-回归任务-1114\03-model\03-models_params-6D_copy.npy'
                                 ,space_pre)

        print(u)
        pre_a = np.column_stack((pre_a, predict_y))

    pre_all = np.column_stack((np.mean(pre_a[:, 1:, ], axis=1), np.std(pre_a[:, 1:, ], axis=1)))

    PA = PA(50, pre_all)
    m01m = MinMaxScaler()
    PAs = np.array(PA)
    PA01 = m01m.fit_transform(PAs.reshape(-1, 1))
    pre_all = np.column_stack((pre_all, PA))
    pre_all = np.column_stack((pre_all, PA01))

    data_write(pre_all, "08-Def_pas.xlsx")
