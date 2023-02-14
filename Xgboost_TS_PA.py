import warnings

import joblib
import numpy as np
import openpyxl
import pandas as pd
from scipy import stats
from sklearn.preprocessing import MinMaxScaler
from sklearn.preprocessing import StandardScaler
from sklearn.utils import resample

warnings.filterwarnings("ignore")
rng = np.random.RandomState(0)
np.set_printoptions(threshold=np.inf)
np.set_printoptions(suppress=True)


def data_read():
    data = pd.read_excel(r'dataset.csv')
    x0 = data.iloc[:, 1:-2]
    y_Tem = data.iloc[:, -2]

    y_Tem = np.ravel(y_Tem)
    y_Tem.astype('float')

    return x0, y_Tem


def model_pre(x, y, name, name_path, para_path, space_pre):
    params = np.load(para_path,allow_pickle=True)
    model = joblib.load(name_path)

    X_ = x
    y_ = y

    mm_y = MinMaxScaler()
    for param in params:
        if param['type'] == name:
            m = param
            if 'normalize' in param:
                if param['normalize'] == 1:
                    mm_x = MinMaxScaler()
                    X_ = mm_x.fit_transform(X_)
                    space_pre = mm_x.fit_transform(space_pre)

            if 'scale' in param:
                if param['scale'] == 1:
                    ss_x = StandardScaler()
                    X_ = ss_x.fit_transform(X_)
                    space_pre = ss_x.fit_transform(space_pre)

            if 'normalize_y' in param:
                if param['normalize_y'] == 1:
                    y_ = mm_y.fit_transform(y_.reshape(-1, 1))
                    y_ = np.ravel(y_)
                    y_.astype(int)

    model.fit(X_, y_)

    y_pre = model.predict(space_pre)

    if m['normalize_y'] == 1:
        y_pre = mm_y.inverse_transform(y_pre.reshape(-1, 1))

    return y_pre


def PA(goal, searchspace0):
    p = (goal - searchspace0[:, 0]) / (searchspace0[:, 1])
    pas = 1 - stats.norm.cdf(p)
    return (pas)


def data_write(data, path):
    outwb = openpyxl.Workbook()
    ws = outwb.create_sheet(index=1)
    i = 1
    r = 1
    for line in data:
        for col in range(1, len(line) + 1):
            # ColNum = r
            ws.cell(row=r, column=col).value = line[col - 1]
        i += 1
        r += 1
    savexlsx = path
    outwb.save(savexlsx)


if __name__ == '__main__':
    stand_scaler = StandardScaler()
    norm_scaler = MinMaxScaler()

    space_pre = pd.read_excel('To_predict.csv')
    x0,  y_Tem = data_read()

    pre_a = np.arange(0, len(space_pre))
    n = len(y_Tem)
    m = 10

    for u in range(0, m):
        x11, y11 = resample(x0, y_Tem, replace=True, n_samples=n, random_state=u)
        predict_y = model_pre(x11, y11
                              , 'xgboost'
                              , r'xgboost_TS.pkl'
                              , r'TS_models_params.npy'
                              , space_pre)
        print(u)
        pre_a = np.column_stack((pre_a, predict_y))

    pre_all = np.column_stack((np.mean(pre_a[:, 1:, ], axis=1)
                               , np.std(pre_a[:, 1:, ], axis=1)))

    PA = PA(350, pre_all)
    m01m = MinMaxScaler()
    PAs = np.array(PA)
    PA01 = m01m.fit_transform(PAs.reshape(-1, 1))
    pre_all = np.column_stack((pre_all, PA))
    pre_all = np.column_stack((pre_all, PA01))

    data_write(pre_all, "TS_pas.csv")
